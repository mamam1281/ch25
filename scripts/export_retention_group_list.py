import argparse
import csv
import datetime as dt
import re
from collections import Counter
from pathlib import Path

import openpyxl


_DATE_RE = re.compile(
    r"^\s*(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?\s*$"
)


def _parse_date(value: object) -> dt.date | None:
    if value is None:
        return None

    if isinstance(value, dt.datetime):
        return value.date()

    if isinstance(value, dt.date):
        return value

    if isinstance(value, (int, float)):
        return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        m = _DATE_RE.match(text)
        if not m:
            return None
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return dt.date(year, month, day)
        except ValueError:
            return None

    return None


def _first_non_empty_row(ws, max_scan_rows: int = 30):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            return row_idx, list(row)
    return 1, []


def _normalize_header(header_row: list[object]) -> list[str]:
    return [str(h).strip() if h is not None else "" for h in header_row]


def _find_sheet(wb, preferred: str) -> str:
    if preferred in wb.sheetnames:
        return preferred

    # Fallback: choose the closest match
    for name in wb.sheetnames:
        if preferred in name:
            return name
    raise SystemExit(f"Sheet not found: {preferred}. Available: {wb.sheetnames}")


def _col_index(header: list[str], candidates: list[str]) -> int | None:
    for c in candidates:
        if c in header:
            return header.index(c)

    # loose match (spaces removed)
    normalized = {h.replace(" ", ""): i for i, h in enumerate(header)}
    for c in candidates:
        key = c.replace(" ", "")
        if key in normalized:
            return normalized[key]

    return None


def export_retention_group_people(
    xlsx_path: Path,
    out_csv: Path,
    out_report_md: Path,
    cutoff_date: dt.date,
    sheet_name: str,
):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    actual_sheet = _find_sheet(wb, sheet_name)
    ws = wb[actual_sheet]

    header_row_idx, header_raw = _first_non_empty_row(ws)
    header = _normalize_header(header_raw)

    idx_group = _col_index(header, ["리텐션그룹(업데이트)", "리텐션그룹(업데이트 )", "리텐션그룹 업데이트"])
    idx_nick = _col_index(header, ["닉네임"])
    idx_nick_site = _col_index(header, ["닉네임(사이트)", "닉네임( 사이트 )"])
    idx_first = _col_index(header, ["최초충전일"])
    idx_last = _col_index(header, ["최근충전일"])

    missing = []
    if idx_group is None:
        missing.append("리텐션그룹(업데이트)")
    if idx_nick is None:
        missing.append("닉네임")
    if idx_nick_site is None:
        missing.append("닉네임(사이트)")

    if missing:
        raise SystemExit(f"Required columns not found in header: {missing}\nDetected header: {header}")

    rows: list[dict[str, object]] = []
    min_last: dt.date | None = None
    max_last: dt.date | None = None

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        group = row[idx_group] if idx_group is not None and idx_group < len(row) else None
        nick = row[idx_nick] if idx_nick is not None and idx_nick < len(row) else None
        nick_site = row[idx_nick_site] if idx_nick_site is not None and idx_nick_site < len(row) else None

        group_text = str(group).strip() if group is not None and str(group).strip() != "" else "(미분류)"
        nick_text = str(nick).strip() if nick is not None and str(nick).strip() != "" else ""
        nick_site_text = str(nick_site).strip() if nick_site is not None and str(nick_site).strip() != "" else ""

        if nick_text == "" and nick_site_text == "":
            continue

        first_date = None
        last_date = None
        if idx_first is not None and idx_first < len(row):
            first_date = _parse_date(row[idx_first])
        if idx_last is not None and idx_last < len(row):
            last_date = _parse_date(row[idx_last])

        if last_date is not None:
            if min_last is None or last_date < min_last:
                min_last = last_date
            if max_last is None or last_date > max_last:
                max_last = last_date

        # cutoff filter (if last_date exists)
        if last_date is not None and last_date > cutoff_date:
            continue

        rows.append(
            {
                "retention_group_updated": group_text,
                "nickname": nick_text,
                "nickname_site": nick_site_text,
                "first_charge_date": first_date.isoformat() if first_date else "",
                "last_charge_date": last_date.isoformat() if last_date else "",
            }
        )

    rows.sort(key=lambda r: (str(r["retention_group_updated"]), str(r["nickname"]), str(r["nickname_site"])))

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "retention_group_updated",
                "nickname",
                "nickname_site",
                "first_charge_date",
                "last_charge_date",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    counts = Counter([str(r["retention_group_updated"]) for r in rows])
    total = len(rows)

    out_report_md.parent.mkdir(parents=True, exist_ok=True)
    coverage_line = ""
    if min_last is not None or max_last is not None:
        coverage_line = f"- (최근충전일 스캔) min={min_last}, max={max_last} (cutoff={cutoff_date})\n"

    lines = []
    lines.append(f"# 리텐션 그룹 인원 리스트 리포트 (엑셀 기준, cutoff={cutoff_date})\n")
    lines.append("## 1) 대상\n")
    lines.append(f"- 파일: {xlsx_path.name}\n")
    lines.append(f"- 시트: {actual_sheet}\n")
    lines.append("- 개인정보: `핸드폰` 등 PII 컬럼은 산출물에서 제외\n")
    if coverage_line:
        lines.append("## 2) 데이터 커버리지\n")
        lines.append(coverage_line)

    lines.append("## 3) 그룹별 인원 수(행 기준)\n")
    lines.append(f"- 총 {total} rows\n")
    for group, cnt in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        lines.append(f"- {group}: {cnt}\n")

    rel_csv = out_csv.as_posix()
    try:
        rel_csv = out_csv.relative_to(Path.cwd()).as_posix()
    except Exception:
        pass

    lines.append("\n## 4) 산출물\n")
    lines.append(f"- CSV: {rel_csv}\n")

    out_report_md.write_text("".join(lines), encoding="utf-8")

    print(f"Wrote CSV: {out_csv}")
    print(f"Wrote report: {out_report_md}")


def main():
    parser = argparse.ArgumentParser(description="Export retention group people list from a given Excel")
    parser.add_argument("--xlsx", type=Path, default=Path("CC1229업데이트_완료.xlsx"))
    parser.add_argument("--sheet", type=str, default="회원요약_그룹재분류")
    parser.add_argument("--cutoff", type=str, default="2025-12-29")
    parser.add_argument(
        "--out-csv",
        type=Path,
        default=Path("docs/06_ops/202601/exports/retention_group_people_20251229.csv"),
    )
    parser.add_argument(
        "--out-report",
        type=Path,
        default=Path("docs/06_ops/202601/excel_retention_group_people_report_20251229_ko.md"),
    )
    args = parser.parse_args()

    cutoff = dt.date.fromisoformat(args.cutoff)
    export_retention_group_people(
        xlsx_path=args.xlsx,
        out_csv=args.out_csv,
        out_report_md=args.out_report,
        cutoff_date=cutoff,
        sheet_name=args.sheet,
    )


if __name__ == "__main__":
    main()
