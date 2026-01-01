import argparse
import datetime as dt
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


_DATE_RE = re.compile(r"^\s*(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?\s*$")


def _parse_datetime(value: object) -> dt.datetime | None:
    if value is None:
        return None

    if isinstance(value, dt.datetime):
        return value

    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)

    if isinstance(value, (int, float)):
        return None

    if isinstance(value, str):
        m = _DATE_RE.match(value)
        if not m:
            return None
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        hour = int(m.group(4) or 0)
        minute = int(m.group(5) or 0)
        second = int(m.group(6) or 0)
        try:
            return dt.datetime(year, month, day, hour, minute, second)
        except ValueError:
            return None

    return None


def _first_non_empty_row(ws, max_scan_rows: int = 10):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            return row_idx, list(row)
    return 1, []


def analyze(path: Path, max_scan_rows: int = 200):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print(f"file: {path}")
    print(f"sheets: {wb.sheetnames}")

    for name in wb.sheetnames:
        ws = wb[name]
        header_row_idx, header = _first_non_empty_row(ws)
        header = [str(h).strip() if h is not None else "" for h in header]

        # Build a min/max datetime per column
        min_dt: dict[int, dt.datetime] = {}
        max_dt: dict[int, dt.datetime] = {}
        non_empty_counts = defaultdict(int)

        # Sample a few rows after header
        samples = []
        for row in ws.iter_rows(
            min_row=header_row_idx + 1,
            max_row=header_row_idx + 1 + 5,
            values_only=True,
        ):
            samples.append(list(row))

        # Scan for datetime-ish columns
        scanned = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + max_scan_rows, values_only=True):
            scanned += 1
            for col_idx, value in enumerate(row, start=1):
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    continue
                non_empty_counts[col_idx] += 1
                parsed = _parse_datetime(value)
                if parsed is None:
                    continue
                if col_idx not in min_dt or parsed < min_dt[col_idx]:
                    min_dt[col_idx] = parsed
                if col_idx not in max_dt or parsed > max_dt[col_idx]:
                    max_dt[col_idx] = parsed

        # Emit summary
        print("\n---")
        print(f"sheet: {name}")
        print(f"header_row: {header_row_idx}")
        if header:
            print(f"columns({len(header)}): {header}")
        else:
            print("columns: (could not detect header)")

        if samples:
            print("sample_rows(5):")
            for r in samples:
                print(r)

        # Datetime-ish columns
        dt_cols = []
        for col_idx in sorted(min_dt.keys() | max_dt.keys()):
            col_name = header[col_idx - 1] if col_idx - 1 < len(header) else f"col_{col_idx}"
            dt_cols.append((col_idx, col_name, min_dt.get(col_idx), max_dt.get(col_idx), non_empty_counts.get(col_idx, 0)))

        if dt_cols:
            print("datetime_like_columns:")
            for col_idx, col_name, mn, mx, cnt in dt_cols:
                print(f"- {col_idx}:{col_name}  min={mn}  max={mx}  non_empty={cnt}  scanned_rows={scanned}")
        else:
            print("datetime_like_columns: (none detected in first scan)")


def main():
    parser = argparse.ArgumentParser(description="Quickly inspect an .xlsx for KPI analysis readiness")
    parser.add_argument("path", type=Path)
    parser.add_argument("--max-scan-rows", type=int, default=200)
    args = parser.parse_args()

    if not args.path.exists():
        raise SystemExit(f"File not found: {args.path}")

    analyze(args.path, max_scan_rows=args.max_scan_rows)


if __name__ == "__main__":
    main()
